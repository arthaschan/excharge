#!/usr/bin/env python3
"""Convert Nature Comm dataset (processed_data.xlsx, 8 sheets) to parquet.
Route B: reproduce Nature paper baseline on real charging data.
"""
import openpyxl, pandas as pd, numpy as np, time, sys, os

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = _ROOT + '/data/real/'
# 原始数据集 processed_data.xlsx 路径：可用环境变量 RAW_XLSX 指定，默认放在
# reproduction/data/raw/processed_data.xlsx（下载来源见 data/README.md）
SRC = os.environ.get('RAW_XLSX', _ROOT + '/data/raw/processed_data.xlsx')

os.makedirs(OUT, exist_ok=True)

COLS = ['id','transaction_id','begin_time','end_time','total_charging_kwh',
        'total_charging_min','current_soc','current_energy_meter_value',
        'chargingv','charginga','out_power','charging_gun_temperature1',
        'charging_gun_temperature2','types','class_judge','label']

t0 = time.time()
wb = openpyxl.load_workbook(SRC, read_only=True)
all_frames = []
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        rows.append(row)
    df = pd.DataFrame(rows, columns=COLS)
    df['owner'] = sn
    print(f'{sn}: {len(df):,} rows', flush=True)
    all_frames.append(df)

df = pd.concat(all_frames, ignore_index=True)
print(f'Total: {len(df):,} rows', flush=True)

# Light dtypes
for c in ['chargingv','charginga','out_power','charging_gun_temperature1',
          'charging_gun_temperature2','current_soc','total_charging_kwh','total_charging_min']:
    df[c] = pd.to_numeric(df[c], errors='coerce')

df.to_parquet(f'{OUT}/all_data.parquet')
print(f'Saved all_data.parquet ({os.path.getsize(f"{OUT}/all_data.parquet")/1e6:.1f} MB) in {time.time()-t0:.0f}s', flush=True)

# Stats
print('\n=== Label consistency within transaction ===', flush=True)
g = df.groupby('transaction_id')['label'].nunique()
print(f'Transactions: {len(g):,}')
print(f'  unique labels per tx == 1: {(g==1).sum():,} ({(g==1).mean()*100:.2f}%)')
print(f'  tx with both labels: {(g>1).sum():,}')

print('\n=== Sequence length distribution ===', flush=True)
seq_len = df.groupby('transaction_id').size()
print(f'  tx count: {len(seq_len):,}')
print(f'  len >= 30: {(seq_len>=30).sum():,} ({(seq_len>=30).mean()*100:.1f}%)')
print(f'  len >= 10: {(seq_len>=10).sum():,}')
print(f'  len percentiles: {seq_len.quantile([0.5,0.9,0.95,0.99]).round(0).to_dict()}')

print('\n=== Battery type distribution ===', flush=True)
print(df['types'].value_counts().to_string())

print('\n=== Label by owner ===', flush=True)
print(df.groupby('owner')['label'].agg(['count','sum']).rename(columns={'count':'total','sum':'fault'}).to_string())

# Save sequence-level summary
seq_df = df.groupby('transaction_id').agg(
    owner=('owner','first'),
    label=('label','first'),
    types=('types','first'),
    n_points=('id','count'),
    begin=('begin_time','first'),
    end=('end_time','last'),
).reset_index()
seq_df.to_parquet(f'{OUT}/sequences.parquet')
print(f'\nSaved sequences.parquet ({len(seq_df):,} sequences)', flush=True)
